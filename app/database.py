from __future__ import annotations

import csv
import io
import json
import sqlite3
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .config import DATA_DIR, DB_PATH
from .profile_docx import generate_profile_docx
from .schema import load_schema

FORM_LABELS = {
    "1": "Đăng ký NVQS",
    "2": "Phúc tra NVQS",
    "3": "Dân quân tự vệ",
    "4": "Dự bị động viên",
    "5": "Sĩ quan dự bị",
}

_RESOLVED_DB_PATH: str | None = None


def get_connection() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(_resolve_db_path())
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with get_connection() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                form_code TEXT NOT NULL DEFAULT '1',
                form_label TEXT NOT NULL DEFAULT 'Đăng ký NVQS',
                full_name TEXT NOT NULL,
                citizen_id_number TEXT NOT NULL,
                phone TEXT,
                hometown TEXT,
                current_residence TEXT,
                created_at TEXT NOT NULL,
                payload_json TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS form_interest_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                form_code TEXT NOT NULL,
                form_label TEXT NOT NULL,
                source TEXT NOT NULL,
                client_ip TEXT,
                user_agent TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        _ensure_column(conn, "submissions", "form_code", "TEXT NOT NULL DEFAULT '1'")
        _ensure_column(conn, "submissions", "form_label", "TEXT NOT NULL DEFAULT 'Đăng ký NVQS'")
        conn.commit()


def save_submission(payload: dict[str, Any], form_code: str) -> int:
    personal = payload.get("personal_basic", {})
    created_at = datetime.now(timezone.utc).isoformat()
    form_label = FORM_LABELS.get(form_code, "")
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO submissions (
                form_code,
                form_label,
                full_name,
                citizen_id_number,
                phone,
                hometown,
                current_residence,
                created_at,
                payload_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                form_code,
                form_label,
                str(personal.get("full_name", "")).strip(),
                str(personal.get("citizen_id_number", "")).strip(),
                str(personal.get("phone", "")).strip(),
                str(personal.get("hometown", "")).strip(),
                str(personal.get("current_residence", "")).strip(),
                created_at,
                json.dumps(payload, ensure_ascii=False),
            ),
        )
        conn.commit()
        return int(cursor.lastrowid)


def save_form_interest(form_code: str, source: str, client_ip: str, user_agent: str) -> int:
    created_at = datetime.now(timezone.utc).isoformat()
    form_label = FORM_LABELS.get(form_code, form_code)
    with get_connection() as conn:
        cursor = conn.execute(
            """
            INSERT INTO form_interest_logs (
                form_code,
                form_label,
                source,
                client_ip,
                user_agent,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (form_code, form_label, source, client_ip, user_agent, created_at),
        )
        conn.commit()
        return int(cursor.lastrowid)


def list_submissions() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, form_code, form_label, full_name, citizen_id_number, phone, hometown,
                   current_residence, created_at, payload_json
            FROM submissions
            ORDER BY id DESC
            """
        ).fetchall()
    return [_deserialize_submission_row(row) for row in rows]


def list_submissions_page(
    page: int = 1,
    page_size: int = 10,
    search: str = "",
) -> dict[str, Any]:
    safe_page_size = max(1, min(int(page_size or 10), 100))
    safe_page = max(1, int(page or 1))
    normalized_search = _normalize_search_text(search)

    if normalized_search:
        rows = list_submissions()
        filtered_rows = [row for row in rows if _submission_search_blob(row).find(normalized_search) != -1]
        total_items = len(filtered_rows)
        effective_page_size = total_items if total_items else safe_page_size
        return {
            "items": filtered_rows,
            "page": 1,
            "page_size": effective_page_size,
            "total_items": total_items,
            "total_pages": 1,
            "search": search,
        }

    with get_connection() as conn:
        total_items = int(conn.execute("SELECT COUNT(*) FROM submissions").fetchone()[0])
        total_pages = max(1, (total_items + safe_page_size - 1) // safe_page_size) if total_items else 1
        safe_page = min(safe_page, total_pages)
        offset = (safe_page - 1) * safe_page_size
        rows = conn.execute(
            """
            SELECT id, form_code, form_label, full_name, citizen_id_number, phone, hometown,
                   current_residence, created_at, payload_json
            FROM submissions
            ORDER BY id DESC
            LIMIT ? OFFSET ?
            """,
            (safe_page_size, offset),
        ).fetchall()

    return {
        "items": [_deserialize_submission_row(row) for row in rows],
        "page": safe_page,
        "page_size": safe_page_size,
        "total_items": total_items,
        "total_pages": total_pages,
        "search": search,
    }


def get_submission(submission_id: int) -> dict[str, Any] | None:
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT id, form_code, form_label, full_name, citizen_id_number, phone, hometown,
                   current_residence, created_at, payload_json
            FROM submissions
            WHERE id = ?
            """,
            (submission_id,),
        ).fetchone()
    if not row:
        return None
    return _deserialize_submission_row(row)


def list_form_interest_logs() -> list[dict[str, Any]]:
    with get_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, form_code, form_label, source, client_ip, user_agent, created_at
            FROM form_interest_logs
            ORDER BY id DESC
            """
        ).fetchall()
    return [dict(row) for row in rows]


def summarize_submissions(items: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    rows = items if items is not None else list_submissions()
    interest_logs = list_form_interest_logs()
    now_date = datetime.now(timezone.utc).date()
    today_count = 0
    today_interest_count = 0
    neighborhood_counter: Counter[str] = Counter()
    birth_year_counter: Counter[str] = Counter()
    ward_counter: Counter[str] = Counter()
    training_counter: Counter[str] = Counter()
    submitted_form_counter: Counter[str] = Counter()
    interest_form_counter: Counter[str] = Counter()

    for item in rows:
        payload = item.get("payload", {})
        personal = payload.get("personal_basic", {})
        created_at = _parse_iso_datetime(item.get("created_at"))
        if created_at and created_at.date() == now_date:
            today_count += 1
        _add_counter_value(submitted_form_counter, item.get("form_label"))
        _add_counter_value(neighborhood_counter, personal.get("neighborhood"))
        _add_counter_value(birth_year_counter, _extract_birth_year(personal.get("date_of_birth")))
        _add_counter_value(ward_counter, personal.get("ward"))
        _add_counter_value(training_counter, personal.get("training_level"), case_insensitive=True)

    for item in interest_logs:
        created_at = _parse_iso_datetime(item.get("created_at"))
        if created_at and created_at.date() == now_date:
            today_interest_count += 1
        _add_counter_value(interest_form_counter, item.get("form_label"))

    return {
        "total_submissions": len(rows),
        "today_submissions": today_count,
        "total_interest_logs": len(interest_logs),
        "today_interest_logs": today_interest_count,
        "unique_citizen_ids": len(
            {str(row.get("citizen_id_number", "")).strip() for row in rows if row.get("citizen_id_number")}
        ),
        "top_neighborhoods": _top_counter_items(neighborhood_counter),
        "top_birth_years": _top_counter_items(birth_year_counter),
        "top_wards": _top_counter_items(ward_counter),
        "top_training_levels": _top_counter_items(training_counter),
        "submitted_forms": _top_counter_items(submitted_form_counter, limit=10),
        "interest_forms": _top_counter_items(interest_form_counter, limit=10),
    }


def export_csv() -> str:
    rows = list_submissions()
    headers, flat_rows = build_export_matrix(rows)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in headers])
    for row in flat_rows:
        writer.writerow([row.get(key, "") for key, _ in headers])
    return buffer.getvalue()


def export_excel() -> bytes:
    rows = list_submissions()
    summary = summarize_submissions(rows)
    headers, flat_rows = build_export_matrix(rows)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Dashboard"
    summary_sheet.append(["Chỉ số", "Giá trị"])
    summary_sheet.append(["Tổng số phiếu", summary["total_submissions"]])
    summary_sheet.append(["Số phiếu hôm nay", summary["today_submissions"]])
    summary_sheet.append(["Số CCCD duy nhất", summary["unique_citizen_ids"]])
    summary_sheet.append(["Lượt chọn phiếu đang phát triển", summary["total_interest_logs"]])
    summary_sheet.append(["Lượt chọn hôm nay", summary["today_interest_logs"]])
    summary_sheet.append([])
    summary_sheet.append(["Loại phiếu đã nộp", "Số lượng"])
    for item in summary["submitted_forms"]:
        summary_sheet.append([item["label"], item["count"]])
    summary_sheet.append([])
    summary_sheet.append(["Loại phiếu đang phát triển được chọn", "Số lượng"])
    for item in summary["interest_forms"]:
        summary_sheet.append([item["label"], item["count"]])
    summary_sheet.append([])
    summary_sheet.append(["Top khu phố", "Số lượng"])
    for item in summary["top_neighborhoods"]:
        summary_sheet.append([item["label"], item["count"]])
    summary_sheet.append([])
    summary_sheet.append(["Top năm sinh", "Số lượng"])
    for item in summary["top_birth_years"]:
        summary_sheet.append([item["label"], item["count"]])
    summary_sheet.append([])
    summary_sheet.append(["Top phường", "Số lượng"])
    for item in summary["top_wards"]:
        summary_sheet.append([item["label"], item["count"]])
    summary_sheet.append([])
    summary_sheet.append(["Top trình độ đào tạo", "Số lượng"])
    for item in summary["top_training_levels"]:
        summary_sheet.append([item["label"], item["count"]])

    data_sheet = workbook.create_sheet("Dữ liệu")
    data_sheet.append([label for _, label in headers])
    for row in flat_rows:
        data_sheet.append([row.get(key, "") for key, _ in headers])

    interest_sheet = workbook.create_sheet("Tracking")
    interest_headers = [
        "ID",
        "Loại phiếu",
        "Mã phiếu",
        "Nguồn",
        "IP",
        "User agent",
        "Thời gian tạo",
    ]
    interest_sheet.append(interest_headers)
    for row in list_form_interest_logs():
        interest_sheet.append(
            [
                row["id"],
                row["form_label"],
                row["form_code"],
                row["source"],
                row["client_ip"],
                row["user_agent"],
                row["created_at"],
            ]
        )

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 48)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_profile_docx(submission_id: int) -> bytes | None:
    submission = get_submission(submission_id)
    if submission is None:
        return None
    return generate_profile_docx(submission)


def build_export_matrix(rows: list[dict[str, Any]]) -> tuple[list[tuple[str, str]], list[dict[str, str]]]:
    schema = load_schema()
    headers: list[tuple[str, str]] = [
        ("id", "ID"),
        ("form_code", "Mã phiếu"),
        ("form_label", "Loại phiếu"),
        ("created_at", "Thời gian tạo"),
    ]

    repeatable_lengths = {
        section["id"]: _max_repeatable_length(rows, section["id"])
        for section in schema.get("sections", [])
        if section.get("repeatable")
    }

    for section in schema.get("sections", []):
        section_id = section.get("id", "")
        section_title = section.get("title", section_id)
        if section.get("repeatable"):
            for index in range(repeatable_lengths.get(section_id, 0)):
                for field in section.get("fields", []):
                    field_id = field.get("id", "")
                    field_label = field.get("label", field_id)
                    headers.append(
                        (
                            f"{section_id}.{index}.{field_id}",
                            f"{section_title} {index + 1} - {field_label}",
                        )
                    )
            continue

        for field in section.get("fields", []):
            field_id = field.get("id", "")
            field_label = field.get("label", field_id)
            headers.append((f"{section_id}.{field_id}", f"{section_title} - {field_label}"))

    flat_rows: list[dict[str, str]] = []
    for row in rows:
        payload = row.get("payload", {})
        flat: dict[str, str] = {
            "id": str(row.get("id", "")),
            "form_code": str(row.get("form_code", "")),
            "form_label": str(row.get("form_label", "")),
            "created_at": str(row.get("created_at", "")),
        }
        for section in schema.get("sections", []):
            section_id = section.get("id", "")
            section_payload = payload.get(section_id)
            if section.get("repeatable"):
                items = section_payload if isinstance(section_payload, list) else []
                for index in range(repeatable_lengths.get(section_id, 0)):
                    item_payload = items[index] if index < len(items) and isinstance(items[index], dict) else {}
                    for field in section.get("fields", []):
                        field_id = field.get("id", "")
                        flat[f"{section_id}.{index}.{field_id}"] = _stringify_value(item_payload.get(field_id))
                continue

            section_payload = section_payload if isinstance(section_payload, dict) else {}
            for field in section.get("fields", []):
                field_id = field.get("id", "")
                flat[f"{section_id}.{field_id}"] = _stringify_value(section_payload.get(field_id))
        flat_rows.append(flat)
    return headers, flat_rows


def _deserialize_submission_row(row: sqlite3.Row) -> dict[str, Any]:
    item = dict(row)
    payload_json = item.pop("payload_json")
    try:
        item["payload"] = json.loads(payload_json)
    except json.JSONDecodeError:
        item["payload"] = {}
    return item


def _max_repeatable_length(rows: list[dict[str, Any]], section_id: str) -> int:
    max_length = 0
    for row in rows:
        payload = row.get("payload", {})
        items = payload.get(section_id)
        if isinstance(items, list):
            max_length = max(max_length, len(items))
    return max_length


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(_stringify_value(item) for item in value if _stringify_value(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _top_counter_items(counter: Counter[str], limit: int = 5) -> list[dict[str, Any]]:
    return [{"label": label, "count": count} for label, count in counter.most_common(limit)]


def _normalize_search_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    normalized = unicodedata.normalize("NFD", text)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def _collect_search_tokens(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        tokens: list[str] = []
        for entry in value:
            tokens.extend(_collect_search_tokens(entry))
        return tokens
    if isinstance(value, dict):
        tokens = []
        for entry in value.values():
            tokens.extend(_collect_search_tokens(entry))
        return tokens
    return [str(value)]


def _submission_search_blob(item: dict[str, Any]) -> str:
    return _normalize_search_text(" ".join(_collect_search_tokens(item)))


def _add_counter_value(counter: Counter[str], raw_value: Any, case_insensitive: bool = False) -> None:
    value = str(raw_value or "").strip()
    if value:
        counter[_normalize_counter_label(value, case_insensitive=case_insensitive)] += 1


def _normalize_counter_label(value: str, case_insensitive: bool = False) -> str:
    if not case_insensitive:
        return value
    normalized = " ".join(value.split()).lower()
    return normalized.capitalize()


def _extract_birth_year(raw_value: Any) -> str:
    value = str(raw_value or "").strip()
    if not value:
        return ""
    if "/" in value:
        parts = value.split("/")
        if len(parts) == 3 and len(parts[-1]) == 4 and parts[-1].isdigit():
            return parts[-1]
    if "-" in value:
        parts = value.split("-")
        if len(parts) == 3:
            if len(parts[0]) == 4 and parts[0].isdigit():
                return parts[0]
            if len(parts[-1]) == 4 and parts[-1].isdigit():
                return parts[-1]
    return value if len(value) == 4 and value.isdigit() else ""


def _parse_iso_datetime(value: Any) -> datetime | None:
    try:
        return datetime.fromisoformat(str(value))
    except Exception:
        return None


def _ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str) -> None:
    existing_columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}
    if column_name not in existing_columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def _resolve_db_path() -> str:
    global _RESOLVED_DB_PATH
    if _RESOLVED_DB_PATH is not None:
        return _RESOLVED_DB_PATH

    runtime_path = str(DB_PATH.with_name("bchqs.runtime.db"))
    candidates = [str(DB_PATH), runtime_path]
    for candidate in candidates:
        if _is_usable_sqlite(candidate):
            _RESOLVED_DB_PATH = candidate
            return candidate

    _RESOLVED_DB_PATH = runtime_path
    return _RESOLVED_DB_PATH


def _is_usable_sqlite(path: str) -> bool:
    from pathlib import Path

    try:
        if not Path(path).exists():
            return True
        conn = sqlite3.connect(path)
        conn.execute("PRAGMA schema_version").fetchall()
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("ROLLBACK")
        conn.close()
        return True
    except sqlite3.Error:
        return False
